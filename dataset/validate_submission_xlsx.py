#!/usr/bin/env python3
"""
Validate submission XLSX per challenge rules.
Exactly 1 sheet named "Ranked Candidates".
Row 1 = header. Rows 2–101 = exactly 100 data rows.
Columns: candidate_id, rank, score, reasoning
"""

import re
import sys
from pathlib import Path
import openpyxl

REQUIRED_HEADER = ["candidate_id", "rank", "score", "reasoning"]
CANDIDATE_ID_PATTERN = re.compile(r"^CAND_[0-9]{7}$")
DATA_ROW_START = 2
EXPECTED_DATA_ROWS = 100


def validate_submission(xlsx_path):
    errors = []
    path = Path(xlsx_path)

    if path.suffix.lower() != ".xlsx":
        errors.append("Filename must use a .xlsx extension.")
        return errors

    if not path.exists():
        errors.append(f"File not found: {xlsx_path}")
        return errors

    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception as e:
        errors.append(f"Failed to open workbook: {e}")
        return errors

    # ✅ Exactly 1 worksheet
    if len(wb.sheetnames) != 1:
        errors.append(f"Workbook must contain exactly 1 worksheet. Found {len(wb.sheetnames)}: {wb.sheetnames}")
        # We can continue validating using the active or first sheet if present
    
    ws_name = wb.sheetnames[0] if wb.sheetnames else None
    # ✅ Worksheet name: Ranked Candidates
    if ws_name != "Ranked Candidates":
        errors.append(f"Worksheet name must be 'Ranked Candidates'. Found: '{ws_name}'")

    if not ws_name:
        return errors

    ws = wb[ws_name]

    # Load all rows into memory
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        errors.append("Worksheet is empty.")
        return errors

    # ✅ Header verification
    header = [cell for cell in rows[0]]
    if header != REQUIRED_HEADER:
        errors.append(
            "Row 1 (header) columns must be exactly:\n"
            f"  {REQUIRED_HEADER}\n"
            f"Found:\n"
            f"  {header}"
        )

    data_rows = rows[1:]
    # Remove trailing empty rows if any
    while data_rows and all(cell is None for cell in data_rows[-1]):
        data_rows.pop()

    # ✅ Exactly 100 candidates
    n = len(data_rows)
    if n != EXPECTED_DATA_ROWS:
        errors.append(
            f"After the header (row 1), there must be exactly {EXPECTED_DATA_ROWS} "
            f"data rows (rows {DATA_ROW_START}–{DATA_ROW_START + EXPECTED_DATA_ROWS - 1}); "
            f"found {n}."
        )

    seen_ids = set()
    seen_ranks = set()
    by_rank = []

    for i, cells in enumerate(data_rows):
        row_num = DATA_ROW_START + i

        if len(cells) < len(REQUIRED_HEADER):
            errors.append(
                f"Row {row_num}: expected {len(REQUIRED_HEADER)} columns, got {len(cells)}."
            )
            continue

        cid = cells[0]
        rank_val = cells[1]
        score_val = cells[2]
        reasoning = cells[3]

        # ✅ candidate_id checks
        if cid is None:
            errors.append(f"Row {row_num}: candidate_id is required.")
        elif not isinstance(cid, str):
            errors.append(f"Row {row_num}: candidate_id must be a string. Got type {type(cid).__name__}.")
        else:
            cid_s = cid.strip()
            if not CANDIDATE_ID_PATTERN.match(cid_s):
                errors.append(
                    f"Row {row_num}: candidate_id must match CAND_XXXXXXX (7 digits). Found: '{cid_s}'"
                )
            elif cid_s in seen_ids:
                errors.append(f"Row {row_num}: duplicate candidate_id '{cid_s}'.")
            else:
                seen_ids.add(cid_s)

        # ✅ rank checks
        if rank_val is None:
            errors.append(f"Row {row_num}: rank is required.")
        elif not isinstance(rank_val, int) or isinstance(rank_val, bool):
            errors.append(f"Row {row_num}: rank must be stored as an integer. Got type {type(rank_val).__name__}.")
        else:
            if not 1 <= rank_val <= 100:
                errors.append(f"Row {row_num}: rank must be between 1 and 100. Found: {rank_val}")
            elif rank_val in seen_ranks:
                errors.append(f"Row {row_num}: duplicate rank {rank_val}.")
            else:
                seen_ranks.add(rank_val)

        # ✅ score checks
        if score_val is None:
            errors.append(f"Row {row_num}: score is required.")
        elif not isinstance(score_val, (int, float)) or isinstance(score_val, bool):
            errors.append(f"Row {row_num}: score must be numeric. Got type {type(score_val).__name__}.")
        else:
            pass

        # Collect for order & tie-break validations
        if cid and rank_val is not None and score_val is not None:
            by_rank.append((rank_val, score_val, cid))

    # ✅ Each rank 1-100 appears exactly once
    missing = set(range(1, 101)) - seen_ranks
    if missing:
        errors.append(
            f"Each rank 1–100 must appear exactly once; missing: {sorted(missing)}"
        )

    # Verify score sort and tie-breaking rules
    by_rank.sort(key=lambda x: x[0])

    # ✅ Scores are in descending order
    for i in range(len(by_rank) - 1):
        r1, s1, _ = by_rank[i]
        r2, s2, _ = by_rank[i + 1]
        if s1 < s2:
            errors.append(
                f"score must be non-increasing by rank: "
                f"rank {r1} ({s1}) < rank {r2} ({s2})."
            )

    # Tie-breaker checks (candidate_id ascending when scores are equal)
    for i in range(len(by_rank) - 1):
        r1, s1, c1 = by_rank[i]
        r2, s2, c2 = by_rank[i + 1]
        if s1 == s2 and c1 > c2:
            errors.append(
                f"Equal scores at ranks {r1} and {r2}: "
                f"tie-break requires candidate_id ascending "
                f"({c1!r} > {c2!r})."
            )

    return errors


def main():
    if len(sys.argv) != 2:
        print("Usage: python validate_submission_xlsx.py <participant_id>.xlsx")
        sys.exit(1)

    errors = validate_submission(sys.argv[1])
    if errors:
        print(f"Validation failed ({len(errors)} issue(s)):\n")
        for e in errors:
            print(f"- {e}")
        sys.exit(1)

    print("Submission is valid.")


if __name__ == "__main__":
    main()
