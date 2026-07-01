import sys
from pathlib import Path

# Add backend directory to PYTHONPATH
root_dir = Path(__file__).resolve().parents[2]
sys.path.append(str(root_dir))
sys.path.append(str(root_dir / "backend"))

from app.schemas.ranking import CandidateRank, RankingResult
from app.schemas.explainability import ExplainabilityReport, ExplanationType, ExplainabilityMetadata
from services.submission_writer import SubmissionWriter
from dataset.validate_submission_xlsx import validate_submission

def test_export_and_validation():
    # 1. Create a dummy ranking result with 105 candidates
    rankings = []
    explanations = {}
    for i in range(105):
        cid = f"CAND_{i:07d}"
        rankings.append(CandidateRank(
            candidate_id=cid,
            job_id="job_abc",
            final_score=max(0.001, 1.0 - (i * 0.005)),
            rank_position=i + 1,
            passed_gates=True,
            failed_dimensions=[]
        ))
        explanations[cid] = ExplainabilityReport(
            candidate_id=cid,
            job_id="job_abc",
            explanation_type=ExplanationType.RECRUITER,
            summary=f"Candidate {cid} shows strong matching capabilities.",
            dimension_explanations={},
            strengths=[],
            weaknesses=[],
            interview_focus=[],
            development_opportunities=[],
            comparisons=[],
            metadata=ExplainabilityMetadata(
                model_version="1.0",
                prompt_version="1.0",
                timestamp="2026-07-01T12:00:00Z",
                explanation_confidence=0.9,
                generation_duration_ms=100
            )
        )

    ranking_result = RankingResult(
        job_id="job_abc",
        rankings=rankings,
        total_ranked=105
    )

    # 2. Run SubmissionWriter
    writer = SubmissionWriter(export_dir=root_dir / "data/submissions")
    xlsx_path = writer.write_csv(ranking_result, explanations, top_n=100, team_id="team_test")

    print(f"XLSX successfully written to: {xlsx_path}")
    assert xlsx_path.exists(), "XLSX file does not exist!"
    assert xlsx_path.suffix == ".xlsx", "Returned path suffix is not .xlsx!"

    csv_path = xlsx_path.with_suffix(".csv")
    assert csv_path.exists(), "CSV file does not exist!"

    # 3. Validate using the new validator
    errors = validate_submission(xlsx_path)
    assert not errors, f"Validation failed: {errors}"
    print("Validation succeeded! File is valid per competition rules.")

    # 4. Check workbook values and types
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    
    # Check worksheet count
    assert len(wb.sheetnames) == 1, f"Expected exactly 1 worksheet, got {len(wb.sheetnames)}"
    
    ws = wb.active
    assert ws.title == "Ranked Candidates", f"Expected worksheet title 'Ranked Candidates', got '{ws.title}'"
    assert ws.freeze_panes == "A2", f"Expected freeze_panes 'A2', got '{ws.freeze_panes}'"

    # Verify header
    headers = [cell.value for cell in ws[1]]
    assert headers == ["candidate_id", "rank", "score", "reasoning"], f"Headers mismatch: {headers}"

    # Verify exactly 100 candidate rows + 1 header row = 101 rows total
    assert ws.max_row == 101, f"Expected 101 rows (1 header + 100 data rows), got {ws.max_row}"

    # Verify ordering of scores & ranks
    scores = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
    assert scores == sorted(scores, reverse=True), "Scores are not sorted in descending order!"

    ranks = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
    assert ranks == list(range(1, 101)), f"Ranks must be exactly 1 to 100 in order, got: {ranks[:10]}"

    # Row 2 (first data row): candidate_id (str), rank (int), score (float), reasoning (str)
    row2_vals = [cell.value for cell in ws[2]]
    print("Row 2 (First Data Row):", row2_vals)
    print("Row 2 types:", [type(cell.value).__name__ for cell in ws[2]])

    assert isinstance(ws[2][0].value, str)
    assert isinstance(ws[2][1].value, int)
    assert isinstance(ws[2][2].value, (int, float))
    assert isinstance(ws[2][3].value, str)

    # Verify score number format
    assert ws[2][2].number_format == "0.000000"

    print("Workbook format checks passed successfully!")

if __name__ == "__main__":
    test_export_and_validation()
