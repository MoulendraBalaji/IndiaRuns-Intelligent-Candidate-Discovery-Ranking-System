import csv
import openpyxl
from openpyxl.utils import get_column_letter
from io import StringIO
from pathlib import Path
from typing import Dict, Optional

from app.schemas.ranking import RankingResult
from app.schemas.explainability import ExplainabilityReport

class SubmissionWriter:
    """
    Responsible for writing the final competition submission CSV and XLSX.
    Adheres strictly to the required submission schema:
    candidate_id, rank, score, reasoning
    """
    
    HEADERS = ["candidate_id", "rank", "score", "reasoning"]

    def __init__(self, export_dir: str | Path = "data/submissions") -> None:
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _build_rows(self, ranking_result: RankingResult, explanations: Dict[str, ExplainabilityReport], top_n: int = 100):
        """
        Builds the rows to be written, returning a list of dicts.
        """
        rows = []
        for rank in ranking_result.rankings[:top_n]:
            reasoning = "Strong candidate matching role requirements."
            if rank.candidate_id in explanations:
                report = explanations[rank.candidate_id]
                reasoning = report.summary.strip()
                
            rows.append({
                "candidate_id": str(rank.candidate_id),
                "rank": int(rank.rank_position),
                "score": float(rank.final_score),
                "reasoning": str(reasoning)
            })
        return rows

    def generate_csv(self, ranking_result: RankingResult, explanations: Dict[str, ExplainabilityReport], top_n: int = 100) -> str:
        """
        Generates the CSV string content for the submission.
        """
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=self.HEADERS)
        writer.writeheader()

        rows = self._build_rows(ranking_result, explanations, top_n)
        for row in rows:
            writer.writerow({
                "candidate_id": row["candidate_id"],
                "rank": row["rank"],
                "score": f"{row['score']:.6f}",
                "reasoning": row["reasoning"]
            })

        return buffer.getvalue()

    def write_xlsx(
        self,
        ranking_result: RankingResult,
        explanations: Dict[str, ExplainabilityReport],
        top_n: int = 100,
        destination: Optional[str | Path] = None,
        team_id: str = "team_submission"
    ) -> Path:
        """
        Writes the submission XLSX file to disk.
        """
        output_path = Path(destination) if destination else self.export_dir / f"{team_id}.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        # Ensure only 1 sheet exists
        ws = wb.active
        ws.title = "Ranked Candidates"

        # Write header
        ws.append(self.HEADERS)

        # Write data rows
        rows = self._build_rows(ranking_result, explanations, top_n)
        for r in rows:
            ws.append([r["candidate_id"], r["rank"], r["score"], r["reasoning"]])

        # Format rank as integer, score as number with 6 decimal places, and enable freeze panes A2
        ws.freeze_panes = "A2"

        # Apply specific formatting to cell values
        # Header starts at row 1, data starts at row 2
        for row_idx in range(2, len(rows) + 2):
            # Column 2 (rank) - integer formatting
            ws.cell(row=row_idx, column=2).number_format = "0"
            # Column 3 (score) - decimal formatting (6 decimal places)
            ws.cell(row=row_idx, column=3).number_format = "0.000000"

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                # If score column, format it to check length with 6 decimal places
                if cell.row > 1 and cell.column == 3 and isinstance(cell.value, (int, float)):
                    val_str = f"{cell.value:.6f}"
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        wb.save(output_path)
        return output_path

    def write_csv(
        self,
        ranking_result: RankingResult,
        explanations: Dict[str, ExplainabilityReport],
        top_n: int = 100,
        destination: Optional[str | Path] = None,
        team_id: str = "team_submission"
    ) -> Path:
        """
        Writes both CSV and XLSX files for backward compatibility, returning the .xlsx path.
        """
        csv_content = self.generate_csv(ranking_result, explanations, top_n)
        
        csv_output_path = Path(destination) if destination else self.export_dir / f"{team_id}.csv"
        # If destination is provided, it might have .csv suffix, let's write it to CSV first
        csv_output_path.parent.mkdir(parents=True, exist_ok=True)
        csv_output_path.write_text(csv_content, encoding="utf-8", newline="")

        # Calculate XLSX destination path
        if destination:
            xlsx_destination = Path(destination).with_suffix(".xlsx")
        else:
            xlsx_destination = self.export_dir / f"{team_id}.xlsx"

        xlsx_path = self.write_xlsx(ranking_result, explanations, top_n, xlsx_destination, team_id)
        return xlsx_path

